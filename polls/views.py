import random
import string
import pandas as pd
import os
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import DispositivosMedicos, Trazabilidad, AlertaVencimiento, Licitacion, LicitacionDato 
from .forms import AlertaVencimientoForm, DispositivoForm
from django.http import FileResponse, HttpResponse
from django.contrib import messages
from django.contrib.messages import get_messages
from django.core.paginator import Paginator
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.utils.timezone import now
from django.http import JsonResponse
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from firebase_admin import auth
import requests
from django.conf import settings
# import firebase_config  # Importamos la configuración de Firebase
import firebase_admin
from .firebase_config import auth
from .firebase_config import firebase
from firebase_admin import credentials
from datetime import datetime
from django.core.serializers.json import DjangoJSONEncoder
import json
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table
from io import BytesIO
import json
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime
from reportlab.lib.pagesizes import landscape, legal
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Definir el diccionario de mapeo en la vista

# Verifica si el usuario pertenece al grupo 'Administrador'

auth = firebase.auth()

ADMIN_EMAIL = "prueba.licitaciones.1@gmail.com"

def is_admin(user):
    return user.groups.filter(name='Administrador').exists()

def login_user(request):
    # Si ya está autenticado, redirigir directamente
    if request.user.is_authenticated:
        return redirect('home' if is_admin(request.user) else 'homevis')

    if request.method == "POST":
        email = request.POST.get('username')
        password = request.POST.get('password')
        next_url = request.POST.get('next') or '/'

        try:
            # 🔐 Autenticación Firebase
            user_firebase = auth.sign_in_with_email_and_password(email, password)

            # 🔍 Buscar o crear usuario en Django
            django_user = User.objects.filter(email=email).first()

            if not django_user:
                django_user = User.objects.create_user(
                    username=email.split("@")[0],
                    email=email,
                    password=User.objects.make_random_password()
                )
                group_name = 'Administrador' if email.lower() == ADMIN_EMAIL else 'Visualizador'
                group, _ = Group.objects.get_or_create(name=group_name)
                django_user.groups.add(group)

            login(request, django_user)

            # ✅ Redirigir según 'next' o grupo
            return redirect(next_url)

        except Exception:
            messages.error(request, "Usuario o contraseña incorrectos.")
            return redirect('login')

    # En GET: capturar 'next' de la URL y pasarlo al formulario
    return render(request, 'inicio.html', {'next': request.GET.get('next', '')})

#def login_user(request):
    if request.method == "POST":
        email = request.POST['username']  # Usamos email en lugar de 'username'
        password = request.POST['password']

        try:
            # 🔹 Autenticar usuario en Firebase
            user_firebase = auth.sign_in_with_email_and_password(email, password)


            # 🔹 Buscar usuario en Django con el mismo email
            django_user, created = User.objects.get_or_create(email=email, defaults={
                'username': email.split('@')[0]
            })
            
            login(request, django_user)

            #if django_user is None:
            #    messages.error(request, "Usuario no registrado en el sistema local.")
            #    return redirect('login')

            # 🔹 Iniciar sesión en Django
            login(request, django_user)

            # 🔹 Redirigir según el grupo del usuario
            if django_user.groups.filter(name='Administrador').exists():
                return redirect('home')
            elif django_user.groups.filter(name='Visualizador').exists():
                return redirect('homevis')
            else:
                return render(request, 'inicio.html', {'error': 'No perteneces a ningún grupo válido.'})

        except Exception as e:
            messages.error(request, "Usuario o contraseña incorrecto.")
            return redirect('login')

    return render(request, 'inicio.html', {})


def password_reset_request(request):
    if request.method == "POST":
        email = request.POST.get("email")

        try:
            api_key = "AIzaSyB56k71mqe2ebmjABA4ckZd81vT8thUJew"  # 🔹 Reemplaza con la API Key encontrada en Firebase Console
            url = f"https://identitytoolkit.googleapis.com/v1/accounts:sendOobCode?key={api_key}"
            
            data = {
                "requestType": "PASSWORD_RESET",
                "email": email
            }
            
            response = requests.post(url, json=data)
            response_data = response.json()

            if response.status_code == 200:
                messages.success(request, "Se ha enviado un correo de recuperación.")
            else:
                error_message = response_data.get("error", {}).get("message", "Error desconocido.")
                messages.error(request, f"Error: {error_message}")

        except Exception as e:
            messages.error(request, f"Ocurrió un error: {e}")

        return redirect("password_reset")

    return render(request, "password_reset.html")

def logout_user(request):
    logout(request)
    return redirect('login')

@login_required
@user_passes_test(is_admin)
def home(request):
    return render(request,'home.html',{'username': request.user.username})

@login_required
def homevis(request):
    return render(request,'homevis.html',{'username': request.user.username})

@login_required
def registrar_usuario(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        email_confirm = request.POST.get('email_confirm')

        if email != email_confirm:
            messages.error(request, "Los correos electrónicos no coinciden.")
        elif User.objects.filter(username=username).exists():
            messages.error(request, "El nombre de usuario ya está registrado.")
        elif User.objects.filter(email=email).exists():
            messages.error(request, "El correo electrónico ya está registrado.")
        else:
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))

            try:
                # 🔗 Crear usuario en Firebase (método correcto de pyrebase)
                firebase_user = auth.create_user_with_email_and_password(email, password)

                # 🔗 Crear usuario en Django
                user = User.objects.create_user(username=username, email=email, password=password)

                # 🔗 Asignar grupo Visualizador
                visualizador_group, _ = Group.objects.get_or_create(name='Visualizador')
                user.groups.add(visualizador_group)

                messages.success(request, f"Usuario registrado correctamente. La contraseña generada es: {password}")

            except Exception as e:
                # Obtener mensaje limpio del error
                error_msg = str(e)
                if hasattr(e, 'args') and len(e.args) > 0:
                    error_msg = e.args[0]
                messages.error(request, f"Error al registrar en Firebase: {error_msg}")

            return redirect('registro')

    return render(request, 'registro.html', {'username': request.user.username})

CAMPOS_EDITABLES = [
    ('nombre_registro_sanitario', 'Nombre Registro Sanitario'),
    ('registro_sanitario', 'Registro Sanitario'),
    ('no_resolucion', 'No. de Resolución'),
    ('fecha_aprobacion', 'Fecha de Aprobación'),
    ('fecha_vencimiento', 'Fecha de Vencimiento'),
    ('expediente', 'Expediente'),
    ('radicado', 'Radicado'),
    ('modalidad', 'Modalidad'),
    ('titular_registro_sanitario', 'Titular del Registro Sanitario'),
    ('direccion_fabricante', 'Dirección del Fabricante'),
    ('importador', 'Importador'),
    ('acondicionador', 'Acondicionador'),
]

def parse_fecha(valor):
    try:
        return datetime.strptime(valor, '%Y-%m-%d').date() if valor else None
    except ValueError:
        return None
    
@login_required
def dispositivos(request):
    # Obtener filtros
    referencia_lh = request.GET.get('referencia_lh', '').strip()
    referencia_fabricante = request.GET.get('referencia_fabricante', '').strip()
    descripcion_espanol = request.GET.get('descripcion_espanol', '').strip()
    descripcion_ingles = request.GET.get('descripcion_ingles', '').strip()
    registro_sanitario = request.GET.get('registro_sanitario', '').strip()
    marca = request.GET.get('marca', '').strip()
    fabricante = request.GET.get('fabricante', '').strip()
    fabricante_actual = request.GET.get('fabricante', '')
    marca_actual = request.GET.get('marca', '')

    # Construir filtros dinámicos
    filtros = {}
    if referencia_lh:
        filtros['referencia_lh__icontains'] = referencia_lh
    if referencia_fabricante:
        filtros['referencia_fabricante__icontains'] = referencia_fabricante
    if descripcion_espanol:
        filtros['descripcion_espanol__icontains'] = descripcion_espanol
    if descripcion_ingles:
        filtros['descripcion_ingles__icontains'] = descripcion_ingles
    if registro_sanitario:
        filtros['registro_sanitario__icontains'] = registro_sanitario
    if marca:
        filtros['marca'] = marca
    if fabricante:
        filtros['fabricante'] = fabricante

    # Consultar dispositivos
    dispositivos_queryset = DispositivosMedicos.objects.filter(**filtros) if filtros else DispositivosMedicos.objects.none()

    # Paginación (100 dispositivos por página)
    paginator = Paginator(dispositivos_queryset, 100)
    page_number = request.GET.get('page')
    dispositivos = paginator.get_page(page_number)

    # Listas únicas para selects
    marcas = DispositivosMedicos.objects.values_list('marca', flat=True).distinct()
    fabricantes = DispositivosMedicos.objects.values_list('fabricante', flat=True).distinct()
    marca_fabricante_map = dict(DispositivosMedicos.objects.values_list('fabricante', 'marca').distinct())

    data = request.session.pop('excel_preview', None)
    columnas = request.session.pop('excel_columnas', None)
    
    columnas_todas = [field.name for field in DispositivosMedicos._meta.fields]
    columnas_ocultas = ['id', 'alerta_vencimiento']
    columnas_visibles = [col for col in columnas_todas if col not in columnas_ocultas]

    return render(request, "dispositivos.html", {
        'username': request.user.username,
        'dispositivos': dispositivos,
        'marcas': marcas,
        'fabricantes': fabricantes,
        'fabricante_actual': fabricante_actual,
        'marca_actual': marca_actual,
        'marca_fabricante_map_json': json.dumps(marca_fabricante_map, cls=DjangoJSONEncoder),
        'referencia_lh': referencia_lh,
        'referencia_fabricante': referencia_fabricante,
        'descripcion_espanol': descripcion_espanol,
        'descripcion_ingles': descripcion_ingles,
        'registro_sanitario': registro_sanitario,
        'marca': marca,
        'fabricante': fabricante,
        'data': data,
        'columnas': columnas,
        'campos_visibles': columnas_visibles
    })

@login_required
def dispositivos_vis(request):
        # Obtener filtros
    referencia_lh = request.GET.get('referencia_lh', '').strip()
    referencia_fabricante = request.GET.get('referencia_fabricante', '').strip()
    descripcion_espanol = request.GET.get('descripcion_espanol', '').strip()
    descripcion_ingles = request.GET.get('descripcion_ingles', '').strip()
    registro_sanitario = request.GET.get('registro_sanitario', '').strip()
    marca = request.GET.get('marca', '').strip()
    fabricante = request.GET.get('fabricante', '').strip()
    fabricante_actual = request.GET.get('fabricante', '')
    marca_actual = request.GET.get('marca', '')

    # Construir filtros dinámicos
    filtros = {}
    if referencia_lh:
        filtros['referencia_lh__icontains'] = referencia_lh
    if referencia_fabricante:
        filtros['referencia_fabricante__icontains'] = referencia_fabricante
    if descripcion_espanol:
        filtros['descripcion_espanol__icontains'] = descripcion_espanol
    if descripcion_ingles:
        filtros['descripcion_ingles__icontains'] = descripcion_ingles
    if registro_sanitario:
        filtros['registro_sanitario__icontains'] = registro_sanitario
    if marca:
        filtros['marca'] = marca
    if fabricante:
        filtros['fabricante'] = fabricante

    # Consultar dispositivos
    dispositivos_queryset = DispositivosMedicos.objects.filter(**filtros) if filtros else DispositivosMedicos.objects.none()

    # Paginación (100 dispositivos por página)
    paginator = Paginator(dispositivos_queryset, 100)
    page_number = request.GET.get('page')
    dispositivos = paginator.get_page(page_number)

    # Listas únicas para selects
    marcas = DispositivosMedicos.objects.values_list('marca', flat=True).distinct()
    fabricantes = DispositivosMedicos.objects.values_list('fabricante', flat=True).distinct()
    marca_fabricante_map = dict(DispositivosMedicos.objects.values_list('fabricante', 'marca').distinct())

    return render(request, "dispositivosvis.html", {
        'username': request.user.username,
        'dispositivos': dispositivos,
        'marcas': marcas,
        'fabricantes': fabricantes,
        'fabricante_actual': fabricante_actual,
        'marca_actual': marca_actual,
        'marca_fabricante_map_json': json.dumps(marca_fabricante_map, cls=DjangoJSONEncoder),
        'referencia_lh': referencia_lh,
        'referencia_fabricante': referencia_fabricante,
        'descripcion_espanol': descripcion_espanol,
        'descripcion_ingles': descripcion_ingles,
        'registro_sanitario': registro_sanitario,
        'marca': marca,
        'fabricante': fabricante,
    })

@login_required
def descargar_dispositivos(request):
    formato = request.GET.get('formato', 'excel')
    columnas_seleccionadas = request.GET.getlist('columnas')

    # Filtros
    filtros = {}
    if referencia_lh := request.GET.get('referencia_lh', '').strip():
        filtros['referencia_lh__icontains'] = referencia_lh
    if referencia_fabricante := request.GET.get('referencia_fabricante', '').strip():
        filtros['referencia_fabricante__icontains'] = referencia_fabricante
    if descripcion_espanol := request.GET.get('descripcion_espanol', '').strip():
        filtros['descripcion_espanol__icontains'] = descripcion_espanol
    if descripcion_ingles := request.GET.get('descripcion_ingles', '').strip():
        filtros['descripcion_ingles__icontains'] = descripcion_ingles
    if registro_sanitario := request.GET.get('registro_sanitario', '').strip():
        filtros['registro_sanitario__icontains'] = registro_sanitario
    if marca := request.GET.get('marca', '').strip():
        filtros['marca'] = marca
    if fabricante := request.GET.get('fabricante', '').strip():
        filtros['fabricante'] = fabricante

    dispositivos = DispositivosMedicos.objects.filter(**filtros)

    # Si no se seleccionaron columnas, usar por defecto un subconjunto visible
    if not columnas_seleccionadas:
        columnas_todas = [field.name for field in DispositivosMedicos._meta.fields]
        columnas_ocultas = ['id', 'alerta_vencimiento', 'modalidad', 'direccion_fabricante', 'presentacion_comercial']
        columnas_seleccionadas = [col for col in columnas_todas if col not in columnas_ocultas]
        columnas_seleccionadas = sorted(set(columnas_seleccionadas), key=lambda x: x)

    datos_filtrados = [
        tuple(getattr(obj, col) if getattr(obj, col) is not None else '' for col in columnas_seleccionadas)
        for obj in dispositivos
    ]

    if formato == 'excel':
        df = pd.DataFrame(datos_filtrados, columns=columnas_seleccionadas)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            sheet_name = 'Dispositivos'
            df.to_excel(writer, sheet_name=sheet_name, startrow=4, index=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            wrap_format = workbook.add_format({'text_wrap': True, 'valign': 'top', 'font_size': 10})
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'top',
                'font_size': 10, 'bg_color': '#000248', 'font_color': 'white'
            })

            for i, column in enumerate(df.columns):
                max_width = max(df[column].astype(str).map(len).max(), len(column))
                worksheet.set_column(i, i, min(max_width + 2, 35), wrap_format)
                worksheet.write(4, i, column.replace("_", " ").title(), header_format)

            worksheet.merge_range('A1:D1', 'Reporte de Dispositivos Filtrados', workbook.add_format({
                'bold': True, 'font_size': 16, 'align': 'left', 'valign': 'vcenter'
            }))
            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
            worksheet.merge_range('A2:D2', f"Fecha de generación: {fecha_actual}", workbook.add_format({
                'font_size': 10, 'align': 'left', 'valign': 'vcenter'
            }))

        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="dispositivos.xlsx"'
        return response

    elif formato == 'pdf':
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(legal),
            rightMargin=10,
            leftMargin=10,
            topMargin=10,
            bottomMargin=10
        )
        elements = []

        # Encabezado del PDF
        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Dispositivos Filtrados", styles['Title'])
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        fecha = Paragraph(f"Fecha de generación: {fecha_actual}", styles['Normal'])
        elements.extend([title, fecha, Spacer(1, 12)])

        # Tabla
        cell_style = ParagraphStyle(name='cell_style', fontSize=6, leading=6, alignment=1)  # Centered
        header_row = [
        Paragraph(f'<b><font color="white">{col.replace("_", " ").title()}</font></b>', cell_style)
        for col in columnas_seleccionadas
]   

        data_rows = [
            [Paragraph(str(cell), cell_style) for cell in fila]
            for fila in datos_filtrados
        ]
        data = [header_row] + data_rows

        ancho_total = 900
        col_width = ancho_total / len(columnas_seleccionadas)
        col_widths = [col_width] * len(columnas_seleccionadas)

        table = Table(data, repeatRows=1, colWidths=col_widths)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        table.setStyle(style)

        elements.append(table)
        doc.build(elements)

        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="dispositivos.pdf"'
        return response

    return HttpResponse("Formato no soportado", status=400)

@csrf_exempt
def get_columns_dispositivos(request):
    if request.method == 'POST' and 'file' in request.FILES:
        file = request.FILES['file']
        extension = os.path.splitext(file.name)[1].lower()

        if extension not in ['.xlsx', '.xls']:
            return JsonResponse({"error": "Solo se permiten archivos con extensión .xlsx o .xls"}, status=400)

        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            # Eliminar filas vacías y contar las válidas
            df = df.dropna(how='all')
            row_count = len(df)

            file_columns = df.columns.tolist()
            db_columns = [field.name for field in DispositivosMedicos._meta.fields if field.name != 'id']

            return JsonResponse({
                "columns": file_columns,
                "db_columns": db_columns,
                "row_count": row_count
            })

        except Exception as e:
            return JsonResponse({"error": f"Error al leer el archivo: {str(e)}"}, status=400)

    return JsonResponse({"error": "No se recibió un archivo válido"}, status=400)


@login_required
def subir_dispositivos_mapeo(request):
    if request.method == 'POST' and 'file' in request.FILES:
        file = request.FILES['file']
        column_mappings = request.POST.get('column_mappings', '{}')
        extension = os.path.splitext(file.name)[1].lower()

        if extension not in ['.xlsx', '.xls']:
            messages.error(request, "Solo se permiten archivos con extensión .xlsx o .xls.")
            return redirect('dispositivos')

        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            column_mappings = json.loads(column_mappings)
            df.rename(columns=column_mappings, inplace=True)

            modelo_campos = {field.name for field in DispositivosMedicos._meta.fields if field.name != 'id'}
            columnas_validas = [col for col in df.columns if col in modelo_campos]

            df = df.dropna(how='all')  # Eliminar filas completamente vacías
            df.drop_duplicates(subset=['referencia_lh', 'referencia_fabricante', 'descripcion_espanol'], inplace=True)  # Eliminar duplicados exactos

            nuevos_dispositivos = []
            for _, fila in df.iterrows():
                datos = {campo: fila[campo] for campo in columnas_validas if campo in fila and pd.notna(fila[campo])}

                for fecha_campo in ['fecha_aprobacion', 'fecha_vencimiento']:
                    if fecha_campo in datos:
                        try:
                            datos[fecha_campo] = pd.to_datetime(datos[fecha_campo], dayfirst=True, errors='coerce')
                        except:
                            datos[fecha_campo] = None

                nuevos_dispositivos.append(DispositivosMedicos(**datos))

            DispositivosMedicos.objects.bulk_create(nuevos_dispositivos)
            messages.success(request, f"Se importaron {len(nuevos_dispositivos)} dispositivos correctamente.")

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect('dispositivos')


@login_required
def agregar_dispositivo_individual(request):
    if request.method == 'POST':
        fabricante = request.POST.get('fabricante', '').strip()
        marca = request.POST.get('marca', '').strip()
        descripcion_espanol = request.POST.get('descripcion_espanol', '').strip()
        descripcion_ingles = request.POST.get('descripcion_ingles', '').strip()
        referencia_lh = request.POST.get('referencia_lh', '').strip()
        referencia_fabricante = request.POST.get('referencia_fabricante', '').strip()
        registro_sanitario = request.POST.get('registro_sanitario', '').strip()
        nombre_registro_sanitario = request.POST.get('nombre_registro_sanitario', '').strip()
        titular_registro_sanitario = request.POST.get('titular_registro_sanitario', '').strip()
        radicado = request.POST.get('radicado', '').strip()
        no_resolucion = request.POST.get('no_resolucion', '').strip()
        expediente = request.POST.get('expediente', '').strip()
        modelo = request.POST.get('modelo', '').strip()
        direccion_fabricante = request.POST.get('direccion_fabricante', '').strip()
        importador = request.POST.get('importador', '').strip()
        acondicionador = request.POST.get('acondicionador', '').strip()
        modalidad = request.POST.get('modalidad', '').strip()
        clasificacion_riesgo = request.POST.get('clasificacion_riesgo', '').strip()
        vida_util_anos = request.POST.get('vida_util_anos', '').strip()
        vida_util_meses = request.POST.get('vida_util_meses', '').strip()
        material_fabricacion = request.POST.get('material_fabricacion', '').strip()
        presentacion_comercial = request.POST.get('presentacion_comercial', '').strip()
        condiciones_almacenamiento = request.POST.get('condiciones_almacenamiento', '').strip()
        tipo_dispositivo = request.POST.get('tipo_dispositivo', '').strip()
        fecha_aprobacion = request.POST.get('fecha_aprobacion', '').strip()
        fecha_vencimiento = request.POST.get('fecha_vencimiento', '').strip()


        if not fabricante or not marca:
            messages.error(request, "Los campos 'Fabricante' y 'Marca' son obligatorios.")
            return redirect('dispositivos')

        nuevo_dispositivo = DispositivosMedicos.objects.create(
            fabricante=fabricante,
            marca=marca,
            descripcion_espanol=descripcion_espanol,
            descripcion_ingles=descripcion_ingles,
            referencia_lh=referencia_lh,
            referencia_fabricante=referencia_fabricante,
            registro_sanitario=registro_sanitario,
            nombre_registro_sanitario=nombre_registro_sanitario,
            titular_registro_sanitario=titular_registro_sanitario,
            radicado=radicado,
            no_resolucion=no_resolucion,
            expediente=expediente,
            modelo=modelo,
            direccion_fabricante=direccion_fabricante,
            importador=importador,
            acondicionador=acondicionador,
            modalidad=modalidad,
            clasificacion_riesgo=clasificacion_riesgo,
            vida_util_anos=vida_util_anos if vida_util_anos else None,
            vida_util_meses=vida_util_meses if vida_util_meses else None,
            material_fabricacion=material_fabricacion,
            presentacion_comercial=presentacion_comercial,
            condiciones_almacenamiento=condiciones_almacenamiento,
            tipo_dispositivo=tipo_dispositivo,
            fecha_aprobacion=parse_fecha(fecha_aprobacion),
            fecha_vencimiento=parse_fecha(fecha_vencimiento)
        )

        messages.success(request, f"✅ Se agregó correctamente la nueva referencia: {referencia_lh}")
        return redirect('dispositivos')

    return redirect('dispositivos')

def eliminar_dispositivo(request, id):
    dispositivo = get_object_or_404(DispositivosMedicos, id=id)
    dispositivo.delete()
    messages.success(request, 'La referencia fue eliminada exitosamente.')
    return redirect('dispositivos')  # Asegúrate que esta sea tu URL correcta

@login_required
def edicion_masiva(request):
    # Capturar filtros
    no_resolucion = request.GET.get('no_resolucion', '').strip()
    registro_sanitario = request.GET.get('registro_sanitario', '').strip()
    marca = request.GET.get('marca', '').strip()
    fabricante = request.GET.get('fabricante', '').strip()
    fabricante_actual = fabricante
    marca_actual = marca

    filtros = {}
    if no_resolucion:
        filtros['no_resolucion__icontains'] = no_resolucion
    if registro_sanitario:
        filtros['registro_sanitario__icontains'] = registro_sanitario
    if marca:
        filtros['marca'] = marca
    if fabricante:
        filtros['fabricante'] = fabricante

    dispositivos = DispositivosMedicos.objects.filter(**filtros) if filtros else DispositivosMedicos.objects.none()
    modo_edicion = filtros and dispositivos.exists()

    if filtros and not dispositivos.exists():
        messages.error(request, "No se encontraron dispositivos que cumplan con los filtros aplicados.")

    marcas = DispositivosMedicos.objects.values_list('marca', flat=True).distinct()
    fabricantes = DispositivosMedicos.objects.values_list('fabricante', flat=True).distinct()
    marca_fabricante_map = dict(DispositivosMedicos.objects.values_list('fabricante', 'marca').distinct())

    valores_actuales = {}
    if dispositivos.exists():
        for campo, _ in CAMPOS_EDITABLES:
            valores_actuales[campo] = dispositivos.values_list(campo, flat=True).distinct()

    return render(request, "edicion_masiva.html", {
        'fabricantes': fabricantes,
        'marcas': marcas,
        'fabricante_actual': fabricante_actual,
        'marca_actual': marca_actual,
        'marca_fabricante_map_json': json.dumps(marca_fabricante_map),
        'no_resolucion': no_resolucion,
        'registro_sanitario': registro_sanitario,
        'marca': marca,
        'fabricante': fabricante,
        'campos_editables': CAMPOS_EDITABLES,
        'valores_actuales': valores_actuales,
        'filtros': filtros,
        'cantidad_dispositivos': dispositivos.count(),
        'modo_edicion': modo_edicion,
        'ids_filtrados': list(dispositivos.values_list('id', flat=True)),  # ⭐ Pasamos IDs al template
        'username': request.user.username
    })


@login_required
def editar_dispositivos(request):
    if request.method == 'POST':
        ids_string = request.POST.get('ids_filtrados', '')
        ids_lista = ids_string.split(',') if ids_string else []

        if not ids_lista:
            messages.error(request, "Error: No se encontraron dispositivos para actualizar.")
            return redirect('edicion_masiva')

        dispositivos = DispositivosMedicos.objects.filter(id__in=ids_lista)
        referencias_afectadas = list(dispositivos.values_list('referencia_lh', flat=True))
        cambios_realizados = 0

        for campo, _ in CAMPOS_EDITABLES:
            nuevo_valor = request.POST.get(campo, '').strip()
            if nuevo_valor:
                if campo in ['fecha_aprobacion', 'fecha_vencimiento']:
                    try:
                        nuevo_valor = datetime.strptime(nuevo_valor, "%Y-%m-%d").date()
                    except ValueError:
                        continue

                # Registrar trazabilidad antes del cambio
                datos_anteriores = list(dispositivos.values_list(campo, flat=True))
                dispositivos.update(**{campo: nuevo_valor})
                cambios_realizados += len(dispositivos)

                # Verificamos si todos tenían el mismo valor anterior
                if all(valor == datos_anteriores[0] for valor in datos_anteriores):
                    dato_anterior = str(datos_anteriores[0])
                else:
                    dato_anterior = "varios"

                # Registrar trazabilidad
                Trazabilidad.objects.create(
                    columna=campo,
                    dato_anterior=dato_anterior,
                    nuevo_dato=str(nuevo_valor),
                    fecha_hora=now(),
                    referencias_afectadas=json.dumps(referencias_afectadas)
                )

        if cambios_realizados == 0:
            messages.warning(request, "No se realizaron cambios porque no se ingresaron valores válidos o no coincidieron registros.")
        else:
            messages.success(request, f"✅ Se realizaron {cambios_realizados} cambios correctamente.")

        return redirect('edicion_masiva')
    else:
        return redirect('edicion_masiva')

    
@csrf_exempt
def get_columns(request):
    if request.method == 'POST' and 'file' in request.FILES:
        file = request.FILES['file']
        extension = os.path.splitext(file.name)[1].lower()

        if extension not in ['.xlsx', '.xls']:
            return JsonResponse({"error": "Solo se permiten archivos con extensión .xlsx o .xls"}, status=400)

        try:
            df = pd.read_excel(file)
            file_columns = df.columns.str.strip().str.lower().str.replace(' ', '_').tolist()

            db_columns = [field.name for field in DispositivosMedicos._meta.fields if field.name not in ('referencia_lh', 'referencia_fabricante')]
            db_columns.insert(0, 'referencia')

            return JsonResponse({"columns": file_columns, "db_columns": db_columns})

        except Exception as e:
            return JsonResponse({"error": f"Error al leer el archivo: {str(e)}"}, status=400)

    return JsonResponse({"error": "No se recibió un archivo válido"}, status=400)


def descargar_formato_licitacion(request):
    ruta_archivo = os.path.join(settings.BASE_DIR, 'polls','static', 'FORMATO_LICITACIONES.xlsx')
    return FileResponse(open(ruta_archivo, 'rb'), as_attachment=True, filename='FORMATO_LICITACIONES.xlsx')

@login_required
def subir_licitacion(request):
    if request.method == 'POST' and 'file' in request.FILES:
        file = request.FILES['file']
        column_mappings = request.POST.get('column_mappings', '{}')  # Obtener el mapeo de columnas
    
        # file = request.FILES['file']
        extension = os.path.splitext(file.name)[1].lower()

        if extension not in ['.xlsx', '.xls']:
            print("Extensión inválida detectada:", extension)  # DEBUG
            messages.error(request, "Solo se permiten archivos con extensión .xlsx o .xls.")
            return render(request, "licitacion.html", {'username': request.user.username})

        try:
            # Leer el archivo Excel
            df = pd.read_excel(file)

            # Normalizar los nombres de las columnas del archivo (sin espacios y en minúscula)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            # Cargar el mapeo de columnas enviado por el usuario
            column_mappings = json.loads(column_mappings)  # Convertir el JSON a diccionario

            # Renombrar las columnas del archivo según el mapeo proporcionado por el usuario
            df.rename(columns=column_mappings, inplace=True)

            # Validar que la columna referencia_lh esté presente
            if 'referencia' not in df.columns:
                messages.error(request, "El archivo debe contener la columna 'referencia'.")
                return render(request, "licitacion.html", {'username': request.user.username})

            # Obtener los nombres de los campos del modelo
            modelo_campos = {field.name for field in DispositivosMedicos._meta.fields}

            # Identificar columnas que coinciden con la base de datos después del mapeo
            columnas_validas = [col for col in df.columns if col in modelo_campos]

            # Procesar filas del archivo Excel
            filled_data = []
            for _, row in df.iterrows():
                referencia_input = row.get('referencia')


                dispositivo = None
                # if referencia_lh:
                #     dispositivo = DispositivosMedicos.objects.filter(referencia_lh=referencia_lh).first()
                if referencia_input:
                    dispositivo = DispositivosMedicos.objects.filter(referencia_lh=referencia_input).first()

                    if not dispositivo:
                        dispositivo = DispositivosMedicos.objects.filter(referencia_fabricante=referencia_input).first()

                    #Si tampoco, intentar con descripcion_ingles (case-insensitive match)
                    if not dispositivo and referencia_input is not None:
                        referencia_str = str(referencia_input).strip()
                        dispositivo = DispositivosMedicos.objects.filter(descripcion_ingles__iexact=referencia_str).first()

                # Crear un diccionario con la fila original
                row_data = row.to_dict()

                # Si el dispositivo existe, llenar los valores vacíos con los de la base de datos
                if dispositivo:
                    for campo in columnas_validas:
                        if not pd.notna(row_data.get(campo)):  # Solo llenar celdas vacías
                            row_data[campo] = getattr(dispositivo, campo, None)

                filled_data.append(row_data)

            # Convertir la lista de datos llenos a un DataFrame
            output_df = pd.DataFrame(filled_data)

            # Exportar el DataFrame a un archivo Excel para descargar
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="licitaciones_actualizadas.xlsx"'

            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name='Licitacion Actualizada')

                # Mejoras al formato
                workbook = writer.book
                worksheet = writer.sheets['Licitacion Actualizada']

                from openpyxl.styles import Font, Alignment, PatternFill

                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                
                for col_cells in worksheet.iter_cols(min_row=1, max_row=1):
                    for cell in col_cells:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                #Autoajustar ancho columnas
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

                # Congelar primera fila
                worksheet.freeze_panes = worksheet['A2']         

            return response

        except Exception as e:
            # Manejar errores y mostrar mensaje al usuario
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            return render(request, "licitacion.html", {'username': request.user.username})


    return render(request, "licitacion.html", {'username': request.user.username})

@login_required
def subir_licitacion_vis(request):
    if request.method == 'POST' and 'file' in request.FILES:
        file = request.FILES['file']
        column_mappings = request.POST.get('column_mappings', '{}')  # Obtener el mapeo de columnas
    
        # file = request.FILES['file']
        extension = os.path.splitext(file.name)[1].lower()

        if extension not in ['.xlsx', '.xls']:
            print("Extensión inválida detectada:", extension)  # DEBUG
            messages.error(request, "Solo se permiten archivos con extensión .xlsx o .xls.")
            return render(request, "licitacionvis.html", {'username': request.user.username})

        try:
            # Leer el archivo Excel
            df = pd.read_excel(file)

            # Normalizar los nombres de las columnas del archivo (sin espacios y en minúscula)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            # Cargar el mapeo de columnas enviado por el usuario
            column_mappings = json.loads(column_mappings)  # Convertir el JSON a diccionario

            # Renombrar las columnas del archivo según el mapeo proporcionado por el usuario
            df.rename(columns=column_mappings, inplace=True)

            # Validar que la columna referencia_lh esté presente
            if 'referencia' not in df.columns:
                messages.error(request, "El archivo debe contener la columna 'referencia'.")
                return render(request, "licitacionvis.html", {'username': request.user.username})

            # Obtener los nombres de los campos del modelo
            modelo_campos = {field.name for field in DispositivosMedicos._meta.fields}

            # Identificar columnas que coinciden con la base de datos después del mapeo
            columnas_validas = [col for col in df.columns if col in modelo_campos]

            # Procesar filas del archivo Excel
            filled_data = []
            for _, row in df.iterrows():
                referencia_input = row.get('referencia')


                dispositivo = None
                if referencia_input:
                    dispositivo = DispositivosMedicos.objects.filter(referencia_lh=referencia_input).first()

                    if not dispositivo:
                        dispositivo = DispositivosMedicos.objects.filter(referencia_fabricante=referencia_input).first()

                    #Si tampoco, intentar con descripcion_ingles (case-insensitive match)
                    # Si tampoco, intentar con descripcion_ingles (case-insensitive match)
                if not dispositivo and referencia_input is not None:
                    referencia_str = str(referencia_input).strip()
                    dispositivo = DispositivosMedicos.objects.filter(descripcion_ingles__iexact=referencia_str).first()

                # Crear un diccionario con la fila original
                row_data = row.to_dict()

                # Si el dispositivo existe, llenar los valores vacíos con los de la base de datos
                if dispositivo:
                    for campo in columnas_validas:
                        if not pd.notna(row_data.get(campo)):  # Solo llenar celdas vacías
                            row_data[campo] = getattr(dispositivo, campo, None)

                filled_data.append(row_data)

            # Convertir la lista de datos llenos a un DataFrame
            output_df = pd.DataFrame(filled_data)

            # Exportar el DataFrame a un archivo Excel para descargar
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="licitaciones_actualizadas.xlsx"'

            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name='Licitacion Actualizada')

                # Mejoras al formato
                workbook = writer.book
                worksheet = writer.sheets['Licitacion Actualizada']

                from openpyxl.styles import Font, Alignment, PatternFill

                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                
                for col_cells in worksheet.iter_cols(min_row=1, max_row=1):
                    for cell in col_cells:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                #Autoajustar ancho columnas
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

                # Congelar primera fila
                worksheet.freeze_panes = worksheet['A2']         

            return response

        except Exception as e:
            # Manejar errores y mostrar mensaje al usuario
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            return render(request, "licitacionvis.html", {'username': request.user.username})

    return render(request, "licitacionvis.html", {'username': request.user.username})

@login_required
def trazabilidad(request):
    trazabilidad_qs = Trazabilidad.objects.all().order_by('-fecha_hora')

    from django.utils import timezone
    from datetime import datetime, timedelta

    filtro_fecha = request.GET.get('fecha', 'todos')
    hoy_local = timezone.localtime(timezone.now()).date()

    if filtro_fecha == 'hoy':
        inicio = timezone.make_aware(datetime.combine(hoy_local, datetime.min.time()))
        fin = timezone.make_aware(datetime.combine(hoy_local, datetime.max.time()))
        trazabilidad_qs = trazabilidad_qs.filter(fecha_hora__range=(inicio, fin))

    elif filtro_fecha == 'ayer':
        ayer = hoy_local - timedelta(days=1)
        inicio = timezone.make_aware(datetime.combine(ayer, datetime.min.time()))
        fin = timezone.make_aware(datetime.combine(ayer, datetime.max.time()))
        trazabilidad_qs = trazabilidad_qs.filter(fecha_hora__range=(inicio, fin))

    elif filtro_fecha == 'semana':
        hace_7_dias = hoy_local - timedelta(days=7)
        inicio = timezone.make_aware(datetime.combine(hace_7_dias, datetime.min.time()))
        fin = timezone.make_aware(datetime.combine(hoy_local, datetime.max.time()))
        trazabilidad_qs = trazabilidad_qs.filter(fecha_hora__range=(inicio, fin))


    paginator = Paginator(trazabilidad_qs, 5) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Procesa solo los elementos de la página actual
    for t in page_obj:
        try:
            claves = json.loads(t.referencias_afectadas)

            dispositivos = DispositivosMedicos.objects.filter(
                Q(referencia_lh__in=claves) |
                Q(referencia_fabricante__in=claves) |
                Q(descripcion_ingles__in=claves)
            )

            referencias = []
            for d in dispositivos:
                valor_actual = getattr(d, t.columna, None)
                if valor_actual is not None and str(valor_actual).strip().lower() == str(t.nuevo_dato).strip().lower():  # Validamos que efectivamente el campo tiene el nuevo valor
                    if d.referencia_lh:
                        referencias.append(d.referencia_lh)
                    elif d.referencia_fabricante:
                        referencias.append(d.referencia_fabricante)
                    elif d.descripcion_ingles:
                        referencias.append(d.descripcion_ingles)
                    else:
                        referencias.append("(sin referencia)")
            t.referencias_lista = referencias

        except Exception as e:
            print(f"Error: {e}")
            t.referencias_lista = ["(Error al leer)"]

    return render(request, 'trazabilidad.html', {
        'username': request.user.username,
        'trazabilidad': page_obj,
        'fecha_seleccionada': filtro_fecha
    })

@login_required
def vencimiento(request):
    form = AlertaVencimientoForm()
    return render(request, "vencimiento.html", {
        'form': form,
        'username': request.user.username
    })

def eventos_vencimiento(request):
    eventos = []

    # Eliminar duplicados por registro sanitario
    registros_unicos = DispositivosMedicos.objects.values('registro_sanitario').distinct()

    for reg in registros_unicos:
        dispositivo = DispositivosMedicos.objects.filter(registro_sanitario=reg['registro_sanitario']).first()
        try:
            fecha_vencimiento = dispositivo.fecha_vencimiento
            if isinstance(fecha_vencimiento, str):
                fecha_vencimiento = datetime.strptime(fecha_vencimiento.strip(), "%Y-%m-%d").date()

            hoy = datetime.today().date()
            delta = (fecha_vencimiento - hoy).days

            # Asignación de color
            if fecha_vencimiento < hoy:
                color = "red"
            elif delta <= 180:
                color = "orange"
            else:
                color = "green"

            eventos.append({
                "title": dispositivo.registro_sanitario,
                "start": fecha_vencimiento.strftime("%Y-%m-%d"),
                "color": color,
                "nombre_registro_sanitario": dispositivo.nombre_registro_sanitario,
                "fabricante": dispositivo.fabricante,
            })

        except Exception as e:
            print(f"❌ Error al procesar {dispositivo.registro_sanitario}: {e}")

    return JsonResponse(eventos, safe=False)
    
def agregar_evento(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            form = AlertaVencimientoForm(data)

            if form.is_valid():
                form.save()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Formato JSON incorrecto'}, status=400)

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)